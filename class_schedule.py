from openpyxl import load_workbook
import time

day = time.strftime("%d", time.localtime())			# 當前日期
date = time.strftime("%Y-%m-%d", time.localtime())
hour_specify = 8									# 指定每天八點傳訊息
wb = load_workbook('202201.xlsx')					# 讀取班表Excel
sheet = wb['10011']									# 選擇工作表sheet

schedule_h2 = {}
schedule_h3 = {}

# 將H3當天上班狀況存入字典 schedule_h3
for row in range(8,11):
	name = sheet.cell(row, column=2)
	status = sheet.cell(row,column=int(day)+2)
	if status.value == None:
		status = '全班'
	elif status.value == 'A':
		status = '早班'
	elif status.value == '休':
		status = '休假'
	schedule_h3[name.value] = status

# 將H2當天上班狀況存入字典 schedule_h2
for row in range(6,8):
	name = sheet.cell(row, column=2)
	status = sheet.cell(row,column=int(day)+2)
	if status.value == None:
		status = '全班'
	elif status.value == 'A':
		status = '早班'
	elif status.value == '休':
		status = '休假'
	schedule_h2[name.value] = status

sentence_h3 = 'H3今日班表：' + str(date) + '\n'
for n in schedule_h3:
	sentence_h3 += str(n) + ' ' + str(schedule_h3[n]) + '\n'

sentence_h2 = 'H2今日班表：' + str(date) + '\n'
for n in schedule_h2:
	sentence_h2 += str(n) + ' ' + str(schedule_h2[n]) + '\n'