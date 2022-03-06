from openpyxl import load_workbook
import datetime as t

today = t.date.today()
timeString = str(t.date.today())
monthString = timeString[5:7]
print(monthString)							# 今天日期
wb = load_workbook(f'2022{monthString}.xlsx')		# 讀取班表Excel
ws = wb.active									# 選擇工作表sheet


# 班表狀況格式化表示 return(str)
def represent(status):
    if status.value == None:
        status = '全班'
    elif status.value == 'A':
        status = '早班'
    elif status.value == '休':
        status = '休假'
    else:
        status = status.value
    return status


# 將當天上班狀況存入字典 return(dic)
def find_schedule(store='', day=today.day):
    schedule = {}
    if store == 'H3':
        for row in range(8, 11):
            name = ws.cell(row, column=2)
            status = ws.cell(row, column=day+2)
            schedule[name.value] = represent(status)
    elif store == 'H2':
        for row in range(6, 8):
            name = ws.cell(row, column=2)
            status = ws.cell(row, column=day+2)
            schedule[name.value] = represent(status)
    return schedule


# 將班表結果存成字串傳回 return(str)
def show_result(store, result={}, d=today.day):
    now = t.date(today.year, today.month, d)
    sentence = str(now) + f' {store} 當日班表：\n'
    for r in result:
        sentence += r + ' ' + result[r] + '\n'
    return sentence


# schedule.py的主函式
# 結合 find_schedule() 和 show_result() 	return(str)
def schedule(store='', day=today.day):
    result = find_schedule(store, day)
    return show_result(store, result, day)


# 查詢小柯晚上是否一人上班
def find_kk(result={}, day=today.day):
    sentence = ''
    if result['小柯'] == '休假':
        sentence = '小柯今天休假喔！'
    elif result['小柯'] == '早班':
        sentence = '小柯今天早班喔！'
    elif result['JOYA'] == '休假' and result['怡君'] == '早班':
        sentence = '小柯今天晚上一個人上班喔'
    elif result['JOYA'] == '早班' and result['怡君'] == '休假':
        sentence = '小柯今天晚上一個人上班喔'
    else:
        sentence = '小柯今天晚上沒有一個人上班'
    return sentence


# 查詢當月個人早班及休假日期
def find_name_schedule(name=None):
    result_rest = "休假："
    result_A = "早班："
    result_spec = ""
    for row in ws.iter_rows(min_row=6, max_row=10, min_col=2, max_col=2):
        for cell in row:
            if cell.value == name:
                for columns in ws.iter_cols(min_col=3, max_col=33, min_row=cell.row, max_row=cell.row):
                    for c in columns:
                        if c.value:
                            if c.value == "休":
                                result_rest += f"{c.column-2} "
                            elif c.value == "A":
                                result_A += f"{c.column-2} "
                            else:
                                result_spec += f"{c.value}：{c.column-2}\n"

    if result_rest == "休假：" and result_A == "早班：":
        return "查無此人"
    else:
        result = f"{name}\n{result_rest}\n{result_A}\n{result_spec}"
        return result
