import openpyxl
import calendar
from datetime import datetime
import re


class New_schedule:
	# 共同屬性
	def __init__(self,inputTime='202203'):
		self.wb = ''
		self.sheet = ''
		self.year = ''
		self.month = ''

		timeRegex = re.compile(r'202[2-9][01][0-9]')
		mo = timeRegex.match(inputTime)
		if mo == None:
			raise ValueError('請輸入正確日期！')
		else:
			self.year = inputTime[0:4]
			self.month = inputTime[4:6]


	# 讀取舊班表表格
	def open_file(self, file_name='202201.xlsx'):
		self.wb = openpyxl.load_workbook(file_name)
		self.sheet = self.wb.worksheets[0]
		print(self.sheet)

	# 將原本班表底色、儲存格清空
	def clear_cell(self):
		cellRange = self.sheet['C4':'AG10'] # 將原本班表底色清空
		for row in cellRange:
			for c in row:
				c.fill = openpyxl.styles.PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
		
		cellRange = self.sheet['C4':'AG10'] # 將原本儲存格內資料清空
		for row in cellRange:
			for c in row:
				c.value = ''

	# 將當月日期、星期填入儲存格
	def set_month(self):
		month_range = calendar.monthrange(int(self.year), int(self.month))	# 指定要新增的年份、月份
		max_days = month_range[1] # 當月總共有幾天
		weekday = month_range[0] # 當月第一天是星期幾

		# 寫入月份
		month_trans_str = {
		1 : '一', 
		2 : '二', 
		3 : '三', 
		4 : '四', 
		5 : '五', 
		6 : '六', 
		7 : '七', 
		8 : '八', 
		9 : '九', 
		10 : '十', 
		11 : '十一', 
		12 : '十二'
		}
		self.sheet['C2'] = month_trans_str[int(self.month)]

		# 寫入當月日期
		cellRange = self.sheet['C4':'AG4']
		n = 1
		for row in cellRange:
			for c in row:
				if n > max_days:
					break
				c.value = n
				n += 1

		# 寫入當月星期
		weekday_trans_str = {
		0 : '一',
		1 : '二',
		2 : '三', 
		3 : '四', 
		4 : '五', 
		5 : '六', 
		6 : '日'
		}
		cellRange = self.sheet['C5':'AG5']
		n = 1
		for row in cellRange:
			for c in row:
				if n > max_days:
					break
				else:
					c.value = weekday_trans_str[weekday]
					n += 1	
					weekday += 1
					if weekday > 6:
						weekday = 0


		# 將星期六日天數的底色轉為黃色
		cellRange = self.sheet['C5':'AG5']
		list_column = []
		for row in cellRange:
			for c in row:
				if c.value == '六' or c.value == '日':
					list_column.append(c.column)

		for c in list_column:
			for r in range(4, 11):
				current_cell = self.sheet.cell(row = r, column = c)
				current_cell.fill = openpyxl.styles.PatternFill(start_color='FFFF66', end_color='FFFF66', fill_type='solid')

	# 儲存檔案
	def new_file(self):
		self.wb.save(f'{self.year}{self.month}.xlsx')



def main():
	input_time = input('請輸入要新增的年月：')
	try:
		w = New_schedule()
	except ValueError as msg:
		print(msg)
	else:
		w.open_file()
		w.clear_cell()
		w.set_month()
		w.new_file()

if __name__ == '__main__':
	main()